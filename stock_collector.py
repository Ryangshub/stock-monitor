#!/usr/bin/env python3
"""
주식 데이터 자동 수집 및 엑셀 기록 프로그램
  - 국내 주식 : pykrx
  - 해외 주식 : yfinance
  - 매시간 정각(또는 설정 간격) 자동 수집
  - 단일 엑셀 파일 안에 '설정' / '관심종목' / 종목별 시트 통합 관리

사용법:
  python stock_collector.py                        # 스케줄러 모드 (기본 경로)
  python stock_collector.py 경로/관심종목.xlsx       # 엑셀 경로 직접 지정
  python stock_collector.py --init                  # 템플릿 엑셀 생성 후 종료
  python stock_collector.py --once                  # 1회 즉시 수집 후 종료
"""

import argparse
import logging
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# pykrx는 import 시점에 KRX 세션 초기화를 시도하며, 환경변수 미설정 시
# "KRX 로그인 실패" 메시지를 print()로 출력한다. 기능 영향은 없지만
# stdout을 일시 차단해 혼란을 방지한다.
try:
    import io as _io_tmp, sys as _sys_tmp
    _buf = _io_tmp.StringIO()
    _sys_tmp.stdout, _orig = _buf, _sys_tmp.stdout
    try:
        from pykrx import stock as krx
    finally:
        _sys_tmp.stdout = _orig
    del _io_tmp, _sys_tmp, _buf, _orig
except ImportError:
    krx = None

try:
    import yfinance as yf
except ImportError:
    yf = None

from apscheduler.schedulers.blocking import BlockingScheduler

# ─── 상수 ────────────────────────────────────────────────────────────────────
WATCHLIST_SHEET = "관심종목"
SETTINGS_SHEET  = "설정"
KOSPI_SHEET     = "KOSPI지수"

# 개별 종목 시트 컬럼
COLUMNS = [
    "수집시각", "종목코드", "종목명",
    "현재가", "거래량", "등락률(%)",
    "한달최고가", "분기최고가", "반기최고가",
]
COL_WIDTHS = [19, 11, 18, 13, 15, 11, 13, 13, 13]

# KOSPI 지수 시트 컬럼 (현재가 → 지수값)
KOSPI_COLUMNS = [
    "수집시각", "지수코드", "지수명",
    "지수값", "거래량", "등락률(%)",
    "한달최고가", "분기최고가", "반기최고가",
]
KOSPI_COL_WIDTHS = [19, 11, 14, 13, 16, 11, 13, 13, 13]

# 헤더 스타일
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="맑은 고딕", size=10)
# 짝수 행 교차 배경
ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
DATA_FONT = Font(name="맑은 고딕", size=10)

# 숫자 포맷
FMT_KRW   = "#,##0"           # 원화 정수
FMT_USD   = "#,##0.00"        # 달러 소수점 2자리
FMT_IDX   = "#,##0.00"        # 지수값 소수점 2자리 (KOSPI 등)
FMT_PCT   = "0.00%"           # 퍼센트 (값은 0.0X 형태로 저장)
FMT_RATE  = '0.00"%"'          # 등락률 — 값 그대로 표시 + % 접미사 (리터럴 %)


# ─── 로깅 ───────────────────────────────────────────────────────────────────
def setup_logging(log_dir: str):
    log_path = os.path.join(log_dir, "stock_collector.log")
    handlers = [logging.StreamHandler(sys.stdout)]
    try:
        handlers.append(logging.FileHandler(log_path, encoding="utf-8"))
    except OSError:
        pass
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=handlers,
    )
    # pykrx 1.x 내부 로거 포맷 버그(logging.info(args, kwargs))로 인한
    # "--- Logging error ---" stderr 출력 억제
    logging.raiseExceptions = False
    logging.getLogger("pykrx").setLevel(logging.CRITICAL)


log = logging.getLogger(__name__)


# ─── 엑셀 템플릿 생성 ────────────────────────────────────────────────────────
def create_template(path: str):
    """관심종목.xlsx 초기 템플릿 생성"""
    wb = openpyxl.Workbook()
    default = wb.active

    # ── 설정 시트 ──────────────────────────────────────────────────────────
    ws_cfg = wb.create_sheet(SETTINGS_SHEET, 0)
    wb.remove(default)

    _header_row(ws_cfg, 1, ["항목", "값"], widths=[16, 12])

    settings_data = [
        ("시작시간", "09:00"),
        ("종료시간", "16:00"),
        ("수집간격(분)", 60),
    ]
    for r, (k, v) in enumerate(settings_data, 2):
        ws_cfg.cell(r, 1, k).font = DATA_FONT
        ws_cfg.cell(r, 2, v).font = DATA_FONT

    # ── 관심종목 시트 ──────────────────────────────────────────────────────
    ws_wl = wb.create_sheet(WATCHLIST_SHEET, 1)
    _header_row(ws_wl, 1, ["종목코드", "종목명"], widths=[14, 22])

    examples = [
        ("005930", "삼성전자"),
        ("000660", "SK하이닉스"),
        ("069500", "KODEX 200"),
        ("AAPL",   "Apple Inc."),
        ("MSFT",   "Microsoft"),
    ]
    for r, (code, name) in enumerate(examples, 2):
        ws_wl.cell(r, 1, code).font = DATA_FONT
        ws_wl.cell(r, 2, name).font = DATA_FONT

    wb.save(path)
    log.info(f"템플릿 생성 완료: {path}")


def _header_row(ws, row: int, headers: list, widths: list):
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row, col, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 18


# ─── 설정/관심종목 읽기 ───────────────────────────────────────────────────────
def read_settings(path: str) -> dict:
    defaults = {"시작시간": "09:00", "종료시간": "16:00", "수집간격(분)": 60}
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if SETTINGS_SHEET not in wb.sheetnames:
            return defaults
        ws = wb[SETTINGS_SHEET]
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                result[str(row[0]).strip()] = row[1]
        wb.close()
        return {**defaults, **result}
    except Exception as e:
        log.warning(f"설정 읽기 실패 (기본값 사용): {e}")
        return defaults


def read_watchlist(path: str) -> list:
    """[{"code": "005930", "name": "삼성전자"}, ...]"""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        if WATCHLIST_SHEET not in wb.sheetnames:
            log.warning(f"'{WATCHLIST_SHEET}' 시트를 찾을 수 없습니다.")
            return []
        ws = wb[WATCHLIST_SHEET]
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            code = str(row[0]).strip() if row[0] else None
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if code and code.lower() not in ("none", "종목코드", ""):
                items.append({"code": code, "name": name})
        wb.close()
        return items
    except Exception as e:
        log.error(f"관심종목 읽기 실패: {e}")
        return []


# ─── 종목 구분 ────────────────────────────────────────────────────────────────
def is_korean(code: str) -> bool:
    return code.strip().isdigit()


import contextlib, io as _io

@contextlib.contextmanager
def _silent():
    """pykrx의 print/root-logger 노이즈를 일시 억제한다."""
    root = logging.getLogger()
    old_level = root.level
    root.setLevel(logging.CRITICAL)
    buf = _io.StringIO()
    old_stdout = sys.stdout
    sys.stdout = buf
    try:
        yield
    finally:
        sys.stdout = old_stdout
        root.setLevel(old_level)


# ─── 국내 주식 수집 (pykrx) ──────────────────────────────────────────────────
def fetch_korean(code: str, name_hint: str = "") -> dict:
    if krx is None:
        raise RuntimeError("pykrx 미설치 — `pip install pykrx`")

    end_dt  = datetime.now()
    # 반기(약 130거래일) 커버하려면 200일치 캘린더 범위
    start_dt = end_dt - timedelta(days=200)
    end_str  = end_dt.strftime("%Y%m%d")
    start_str = start_dt.strftime("%Y%m%d")

    with _silent():
        df = krx.get_market_ohlcv(start_str, end_str, code)
    if df is None or df.empty:
        raise ValueError(f"[{code}] pykrx 데이터 없음")

    # 종목명: 힌트 우선, API로 보완 (ETF 등 일부는 API 반환 None)
    stock_name = name_hint or code
    try:
        with _silent():
            api_name = krx.get_market_ticker_name(code)
        if api_name:
            stock_name = api_name
    except Exception:
        pass

    latest       = df.iloc[-1]
    current_price = int(latest["종가"])
    volume        = int(latest["거래량"])

    if len(df) >= 2:
        prev_close = float(df.iloc[-2]["종가"])
        change_rate = round((current_price - prev_close) / prev_close * 100, 2) if prev_close else 0.0
    else:
        change_rate = 0.0

    idx = df.index

    def high_since(days: int) -> int:
        cutoff = pd.Timestamp.now() - pd.Timedelta(days=days)
        sub = df[idx >= cutoff]
        return int(sub["고가"].max()) if not sub.empty else current_price

    return {
        "code":        code,
        "name":        stock_name,
        "price":       current_price,
        "volume":      volume,
        "change_rate": change_rate,
        "high_1m":     high_since(30),
        "high_3m":     high_since(90),
        "high_6m":     high_since(180),
        "is_krw":      True,
    }


# ─── 해외 주식 수집 (yfinance) ────────────────────────────────────────────────
def fetch_foreign(ticker: str, name_hint: str = "") -> dict:
    if yf is None:
        raise RuntimeError("yfinance 미설치 — `pip install yfinance`")

    t    = yf.Ticker(ticker)
    hist = t.history(period="6mo", auto_adjust=True)

    if hist.empty:
        raise ValueError(f"[{ticker}] yfinance 데이터 없음")

    # 종목명
    try:
        info = t.info
        name = info.get("longName") or info.get("shortName") or name_hint or ticker
    except Exception:
        name = name_hint or ticker

    latest        = hist.iloc[-1]
    current_price = round(float(latest["Close"]), 4)
    volume        = int(latest["Volume"])

    if len(hist) >= 2:
        prev_close = float(hist.iloc[-2]["Close"])
        change_rate = round((current_price - prev_close) / prev_close * 100, 2) if prev_close else 0.0
    else:
        change_rate = 0.0

    tz = hist.index.tz

    def high_since(days: int) -> float:
        cutoff = pd.Timestamp.now(tz=tz) - pd.Timedelta(days=days)
        sub = hist[hist.index >= cutoff]
        return round(float(sub["High"].max()), 4) if not sub.empty else current_price

    return {
        "code":        ticker,
        "name":        name,
        "price":       current_price,
        "volume":      volume,
        "change_rate": change_rate,
        "high_1m":     high_since(30),
        "high_3m":     high_since(90),
        "high_6m":     high_since(180),
        "is_krw":      False,
    }


# ─── 엑셀 시트 관리 ──────────────────────────────────────────────────────────
def _safe_sheet_name(code: str, name: str) -> str:
    """openpyxl 시트명 규칙: 31자 이하, 특수문자 불가"""
    raw = f"{code}_{name}"
    for ch in r'\/:*?"<>|[]':
        raw = raw.replace(ch, "_")
    return raw[:31]


def ensure_stock_sheet(wb: openpyxl.Workbook, sheet_name: str,
                        columns=None, col_widths=None):
    """종목 또는 지수 데이터 시트를 생성(없을 때만)하고 반환한다."""
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]

    cols   = columns   or COLUMNS
    widths = col_widths or COL_WIDTHS

    ws = wb.create_sheet(sheet_name)
    for col_idx, (header, width) in enumerate(zip(cols, widths), 1):
        cell = ws.cell(1, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 18
    ws.freeze_panes = "A2"
    return ws


# ─── KOSPI 지수 수집 ──────────────────────────────────────────────────────────
def fetch_kospi() -> dict:
    """KOSPI 종합지수 수집.
    1차: yfinance ^KS11  (pykrx 지수 API가 인증 없이는 빈 응답을 반환하므로)
    2차: pykrx get_index_ohlcv_by_date (fallback)
    """
    # ── 1차: yfinance ────────────────────────────────────────────────────────
    if yf is not None:
        try:
            t    = yf.Ticker("^KS11")
            hist = t.history(period="6mo", auto_adjust=True)
            if not hist.empty:
                latest    = hist.iloc[-1]
                cur_val   = round(float(latest["Close"]), 2)
                volume    = int(latest.get("Volume", 0))

                if len(hist) >= 2:
                    prev = float(hist.iloc[-2]["Close"])
                    change_rate = round((cur_val - prev) / prev * 100, 2) if prev else 0.0
                else:
                    change_rate = 0.0

                tz = hist.index.tz
                now_tz = pd.Timestamp.now(tz=tz)
                def hi_yf(days: int) -> float:
                    cutoff = now_tz - pd.Timedelta(days=days)
                    sub = hist[hist.index >= cutoff]
                    return round(float(sub["High"].max()), 2) if not sub.empty else cur_val

                return {
                    "code":        "KOSPI",
                    "name":        "KOSPI 종합지수",
                    "price":       cur_val,
                    "volume":      volume,
                    "change_rate": change_rate,
                    "high_1m":     hi_yf(30),
                    "high_3m":     hi_yf(90),
                    "high_6m":     hi_yf(180),
                }
        except Exception as e:
            log.debug(f"yfinance KOSPI 실패, pykrx로 재시도: {e}")

    # ── 2차: pykrx (fallback) ────────────────────────────────────────────────
    if krx is None:
        raise RuntimeError("pykrx, yfinance 모두 미설치")

    end_dt    = datetime.now()
    start_dt  = end_dt - timedelta(days=200)
    end_str   = end_dt.strftime("%Y%m%d")
    start_str = start_dt.strftime("%Y%m%d")

    with _silent():
        df = krx.get_index_ohlcv_by_date(start_str, end_str, "1001",
                                          name_display=False)
    if df is None or df.empty:
        raise ValueError("KOSPI 지수 데이터 없음 (yfinance·pykrx 모두 실패)")

    latest  = df.iloc[-1]
    cur_val = round(float(latest["종가"]), 2)
    vol_col = next((c for c in ("거래량", "거래대금") if c in df.columns), None)
    volume  = int(latest[vol_col]) if vol_col else 0

    if len(df) >= 2:
        prev = float(df.iloc[-2]["종가"])
        change_rate = round((cur_val - prev) / prev * 100, 2) if prev else 0.0
    else:
        change_rate = 0.0

    now = pd.Timestamp.now()
    def hi_krx(days: int) -> float:
        cutoff = now - pd.Timedelta(days=days)
        sub = df[df.index >= cutoff]
        return round(float(sub["고가"].max()), 2) if not sub.empty else cur_val

    return {
        "code":        "1001",
        "name":        "KOSPI 종합지수",
        "price":       cur_val,
        "volume":      volume,
        "change_rate": change_rate,
        "high_1m":     hi_krx(30),
        "high_3m":     hi_krx(90),
        "high_6m":     hi_krx(180),
    }


# ─── 차트 생성/갱신 ───────────────────────────────────────────────────────────
def update_chart(ws, price_col: int = 4, num_fmt: str = FMT_KRW):
    """K1 위치에 최근 30일 가격 추이 라인 차트를 생성/갱신한다.

    매 수집 시 호출되어 차트를 최신 데이터로 교체한다.
    데이터가 2행 미만이면 건너뛴다.
    """
    cutoff = datetime.now() - timedelta(days=30)

    # 30일 이내 행 인덱스 수집
    data_rows = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row_idx, 1).value
        if val:
            try:
                dt = datetime.strptime(str(val)[:16], "%Y-%m-%d %H:%M")
                if dt >= cutoff:
                    data_rows.append(row_idx)
            except (ValueError, TypeError):
                pass

    if len(data_rows) < 2:
        return

    min_row, max_row = data_rows[0], data_rows[-1]

    # 기존 차트 모두 제거 후 재생성
    ws._charts.clear()

    chart = LineChart()
    chart.title       = "최근 1개월"
    chart.style       = 10          # 흰 배경 깔끔한 스타일
    chart.width       = 22          # cm
    chart.height      = 14          # cm
    chart.y_axis.numFmt   = num_fmt
    chart.y_axis.crossAx  = 100
    chart.x_axis.crossAx  = 200
    chart.x_axis.tickLblSkip = max(1, (max_row - min_row) // 10)  # 레이블 과밀 방지

    # 가격 시리즈 (컬럼 4 = 현재가 / 지수값)
    price_ref = Reference(ws, min_col=price_col, max_col=price_col,
                           min_row=min_row, max_row=max_row)
    chart.add_data(price_ref)

    # 시리즈 스타일: 진한 파란 실선, 스무딩
    s = chart.series[0]
    s.graphicalProperties.line.solidFill = "1F4E79"
    s.graphicalProperties.line.width     = 20000   # 약 1.6 pt
    s.smooth = True

    # X축 카테고리 (수집시각 문자열)
    cats_ref = Reference(ws, min_col=1, max_col=1,
                          min_row=min_row, max_row=max_row)
    chart.set_categories(cats_ref)

    ws.add_chart(chart, "K1")


def write_kospi_row(wb: openpyxl.Workbook, data: dict, collected_at: datetime):
    """KOSPI지수 시트에 1행 추가."""
    ws       = ensure_stock_sheet(wb, KOSPI_SHEET,
                                  columns=KOSPI_COLUMNS,
                                  col_widths=KOSPI_COL_WIDTHS)
    next_row = ws.max_row + 1

    row_values = [
        collected_at.strftime("%Y-%m-%d %H:%M"),
        data["code"],
        data["name"],
        data["price"],
        data["volume"],
        data["change_rate"],
        data["high_1m"],
        data["high_3m"],
        data["high_6m"],
    ]
    alignments = ["center", "center", "left",
                  "right",  "right",  "right",
                  "right",  "right",  "right"]
    formats    = [None, None, None,
                  FMT_IDX, FMT_KRW, FMT_RATE,
                  FMT_IDX, FMT_IDX, FMT_IDX]

    use_alt = (next_row % 2 == 0)
    for col_idx, (val, align, fmt) in enumerate(zip(row_values, alignments, formats), 1):
        cell = ws.cell(next_row, col_idx, val)
        cell.font      = DATA_FONT
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fmt:
            cell.number_format = fmt
        if use_alt:
            cell.fill = ALT_FILL

    update_chart(ws, num_fmt=FMT_IDX)


def write_row(wb: openpyxl.Workbook, data: dict, collected_at: datetime):
    sheet_name = _safe_sheet_name(data["code"], data["name"])
    ws         = ensure_stock_sheet(wb, sheet_name)
    next_row   = ws.max_row + 1
    is_krw     = data.get("is_krw", True)

    price_fmt  = FMT_KRW if is_krw else FMT_USD
    high_fmt   = FMT_KRW if is_krw else FMT_USD

    row_values = [
        collected_at.strftime("%Y-%m-%d %H:%M"),
        data["code"],
        data["name"],
        data["price"],
        data["volume"],
        data["change_rate"],
        data["high_1m"],
        data["high_3m"],
        data["high_6m"],
    ]
    # col별 정렬·포맷 설정
    alignments = ["center", "center", "left",
                  "right", "right", "right",
                  "right", "right", "right"]
    formats = [None, None, None,
               price_fmt, FMT_KRW, FMT_RATE,
               high_fmt, high_fmt, high_fmt]

    use_alt = (next_row % 2 == 0)
    for col_idx, (val, align, fmt) in enumerate(zip(row_values, alignments, formats), 1):
        cell = ws.cell(next_row, col_idx, val)
        cell.font      = DATA_FONT
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fmt:
            cell.number_format = fmt
        if use_alt:
            cell.fill = ALT_FILL

    price_fmt = FMT_KRW if data.get("is_krw", True) else FMT_USD
    update_chart(ws, num_fmt=price_fmt)


# ─── 1회 수집 ────────────────────────────────────────────────────────────────
def collect_once(excel_path: str):
    now = datetime.now()
    log.info(f"{'─'*50}")
    log.info(f"[수집 시작] {now.strftime('%Y-%m-%d %H:%M:%S')}")

    watchlist = read_watchlist(excel_path)
    if not watchlist:
        log.warning("관심종목이 없습니다. 엑셀의 '관심종목' 시트를 확인하세요.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
    except PermissionError:
        log.error("엑셀 파일이 열려 있습니다. 파일을 닫고 다시 시도하세요.")
        return
    except Exception as e:
        log.error(f"엑셀 파일 열기 실패: {e}")
        return

    ok, fail = 0, 0

    # ── KOSPI 지수 수집 ────────────────────────────────────────────────────
    try:
        kospi_data = fetch_kospi()
        write_kospi_row(wb, kospi_data, now)
        log.info(
            f"  ✓ KOSPI 지수 | "
            f"지수값: {kospi_data['price']:,.2f} | "
            f"등락: {kospi_data['change_rate']:+.2f}%"
        )
        ok += 1
    except Exception as e:
        log.error(f"  ✗ KOSPI 지수 수집 실패: {e}")
        fail += 1

    # ── 관심종목 수집 ──────────────────────────────────────────────────────
    for item in watchlist:
        code, name = item["code"], item["name"]
        try:
            if is_korean(code):
                data = fetch_korean(code, name)
            else:
                data = fetch_foreign(code, name)

            # 관심종목 시트에 이름이 있으면 덮어쓰지 않음 (API 값 우선)
            if not data["name"] and name:
                data["name"] = name

            write_row(wb, data, now)
            log.info(
                f"  ✓ {data['name']}({code}) | "
                f"현재가: {data['price']:,.0f} | "
                f"등락: {data['change_rate']:+.2f}%"
            )
            ok += 1
        except Exception as e:
            log.error(f"  ✗ {code} ({name}) 수집 실패: {e}")
            fail += 1

    try:
        wb.save(excel_path)
        log.info(f"[수집 완료] 성공 {ok}건 / 실패 {fail}건 → {excel_path}")
    except PermissionError:
        log.error("저장 실패: 엑셀 파일이 열려 있습니다. 파일을 닫아 주세요.")
    except Exception as e:
        log.error(f"저장 오류: {e}")


# ─── 스케줄러 ────────────────────────────────────────────────────────────────
def _parse_hhmm(s: str) -> tuple:
    """'09:00' → (9, 0)"""
    h, m = str(s).strip().split(":")
    return int(h), int(m)


def _minute_spec(interval: int) -> str:
    """간격(분) → cron minute 문자열. 예) 60→'0', 30→'0,30', 15→'0,15,30,45'"""
    if interval <= 0 or 60 % interval != 0:
        log.warning(f"수집간격 {interval}분은 60의 약수가 아닙니다. 60분으로 대체합니다.")
        return "0"
    minutes = [str(i * interval) for i in range(60 // interval)]
    return ",".join(minutes)


def run_scheduler(excel_path: str):
    settings = read_settings(excel_path)
    start_time = str(settings.get("시작시간", "09:00"))
    end_time   = str(settings.get("종료시간", "16:00"))
    interval   = int(settings.get("수집간격(분)", 60))

    start_h, start_m = _parse_hhmm(start_time)
    end_h,   end_m   = _parse_hhmm(end_time)

    minute_spec = _minute_spec(interval)

    log.info(f"스케줄러 시작 | 운영시간: {start_time} ~ {end_time} | 수집간격: {interval}분")
    log.info(f"엑셀 파일: {excel_path}")
    log.info("Ctrl+C 로 종료")

    def job():
        # 운영시간 내에서만 실행 (APScheduler hour range 보조 검증)
        now = datetime.now()
        start_bound = now.replace(hour=start_h, minute=start_m, second=0, microsecond=0)
        end_bound   = now.replace(hour=end_h,   minute=end_m,   second=59, microsecond=0)
        if start_bound <= now <= end_bound:
            collect_once(excel_path)
        else:
            log.debug(f"운영시간 외 ({now.strftime('%H:%M')}) — 건너뜀")

    scheduler = BlockingScheduler(timezone="Asia/Seoul")
    scheduler.add_job(
        job,
        "cron",
        hour=f"{start_h}-{end_h}",
        minute=minute_spec,
        id="stock_collect",
        misfire_grace_time=300,   # 5분 내 지연 허용
    )

    try:
        scheduler.start()
    except (KeyboardInterrupt, SystemExit):
        log.info("스케줄러 종료")


# ─── CLI ─────────────────────────────────────────────────────────────────────
def main():
    default_path = str(
        Path.home() / "Documents" / "stock_monitor" / "관심종목.xlsx"
    )

    parser = argparse.ArgumentParser(
        description="주식 데이터 자동 수집 및 엑셀 기록",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
예시:
  python stock_collector.py                        # 스케줄러 모드 (기본 경로)
  python stock_collector.py D:/stock/watchlist.xlsx  # 엑셀 경로 직접 지정
  python stock_collector.py --init                  # 템플릿 생성 후 종료
  python stock_collector.py --once                  # 1회 즉시 수집 후 종료
        """,
    )
    parser.add_argument(
        "excel",
        nargs="?",
        default=default_path,
        help=f"엑셀 파일 경로 (기본: {default_path})",
    )
    parser.add_argument(
        "--init",
        action="store_true",
        help="템플릿 엑셀 파일 생성(또는 재생성) 후 종료",
    )
    parser.add_argument(
        "--once",
        action="store_true",
        help="1회 즉시 수집 후 종료 (스케줄 없이)",
    )
    args = parser.parse_args()

    excel_path = args.excel
    log_dir    = os.path.dirname(os.path.abspath(excel_path))

    os.makedirs(log_dir, exist_ok=True)
    setup_logging(log_dir)

    # 미설치 경고
    if krx is None:
        log.warning("pykrx 미설치 — 국내주식 수집 불가. `pip install pykrx`")
    if yf is None:
        log.warning("yfinance 미설치 — 해외주식 수집 불가. `pip install yfinance`")

    # 템플릿 생성 모드
    if args.init:
        if os.path.exists(excel_path):
            ans = input(f"'{excel_path}' 이 이미 존재합니다. 덮어쓰겠습니까? [y/N] ").strip().lower()
            if ans != "y":
                print("취소되었습니다.")
                return
        create_template(excel_path)
        print(f"\n템플릿 생성 완료: {excel_path}")
        print("  1. '관심종목' 시트에 종목코드와 종목명을 입력하세요.")
        print("     국내: 6자리 숫자 코드  |  해외: 영문 티커 (예: AAPL, MSFT)")
        print("  2. '설정' 시트에서 시작/종료 시간과 수집 간격을 조정하세요.")
        print("  3. python stock_collector.py  로 스케줄러를 실행하세요.")
        return

    # 파일 없으면 자동 생성
    if not os.path.exists(excel_path):
        log.info(f"엑셀 파일이 없어 템플릿을 생성합니다: {excel_path}")
        create_template(excel_path)
        print(f"\n템플릿을 생성했습니다: {excel_path}")
        print("  '관심종목' 시트에 종목을 입력한 뒤 다시 실행하세요.")
        return

    if args.once:
        collect_once(excel_path)
    else:
        run_scheduler(excel_path)


if __name__ == "__main__":
    main()
